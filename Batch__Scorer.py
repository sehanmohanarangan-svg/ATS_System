"""
Batch Resume Scorer
====================
Compare a folder of parsed resume JSONs against a single job posting.
Outputs a ranked Excel file with scores, grades, and gap analysis.

Usage:
    python batch_match.py resumes/ job_posting.pdf
    python batch_match.py resumes/ job_posting.pdf --output results.xlsx
    python batch_match.py resumes/ job_posting.pdf --workers 4
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import time
from pathlib import Path
from typing import Any
from concurrent.futures import ThreadPoolExecutor, as_completed

import numpy as np
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# Import your existing modules
from SemanticEngine import HDCSearch
from resume_matcher import (
    TextReader, JobParser,
    score_skills, score_experience, score_education,
    score_summary, score_certifications, score_projects,
    WEIGHTS, _grade, _recommendation,
)


# ---------------------------------------------------------------------------
# Batch scorer
# ---------------------------------------------------------------------------

class BatchScorer:
    def __init__(self) -> None:
        print("Loading semantic model...", flush=True)
        self._sem     = HDCSearch()
        self._reader  = TextReader()
        self._jparser = JobParser()
        print("Model ready.", flush=True)

    def load_job(self, job_path: Path) -> dict[str, Any]:
        """Parse and pre-encode the job posting once for all resumes."""
        raw   = self._reader.read(job_path)
        clean = self._reader.clean(raw)
        job   = self._jparser.parse(clean)

        # Pre-encode job texts into cache
        desc = job.get("description") or job.get("full_text") or ""
        if desc:
            self._sem.encode(desc[:1500])

        job_skills = job.get("required_skills") or job.get("all_skills") or []
        if job_skills:
            self._sem.encode_batch(job_skills)

        print(f"Job: {job.get('title') or job_path.name}", flush=True)
        return job

    def score_one(self, resume: dict, job: dict) -> dict[str, Any]:
        """Score a single resume against the pre-loaded job."""
        # Warm cache for this resume
        self._sem.warm_batch(resume)

        sections = {
            "skills":         score_skills(resume, job, self._sem),
            "experience":     score_experience(resume, job, self._sem),
            "education":      score_education(resume, job, self._sem),
            "summary":        score_summary(resume, job, self._sem),
            "certifications": score_certifications(resume, job, self._sem),
            "projects":       score_projects(resume, job, self._sem),
        }

        total = sum(s.weighted for s in sections.values())

        matched_skills = sections["skills"].matched[:5]
        missing_skills = sections["skills"].missing[:5]
        gaps = [
            f"{k} ({WEIGHTS[k]}pts)"
            for k, s in sections.items()
            if s.raw < 0.50 and WEIGHTS[k] >= 10
        ]
        for s in sections.values():
            gaps += s.missing[:2]

        derived  = resume.get("derived") or {}
        contact  = resume.get("contact") or {}

        return {
            # Identity
            "name":             resume.get("name") or "Unknown",
            "email":            contact.get("email") or "",
            "phone":            contact.get("phone") or "",
            "location":         contact.get("location") or "",
            "source_file":      resume.get("source_file") or "",
            # Overall
            "total_score":      round(total, 1),
            "grade":            _grade(total),
            "recommendation":   _recommendation(total, sections),
            # Section scores (weighted points)
            "skills_score":     round(sections["skills"].weighted, 1),
            "experience_score": round(sections["experience"].weighted, 1),
            "education_score":  round(sections["education"].weighted, 1),
            "summary_score":    round(sections["summary"].weighted, 1),
            "certs_score":      round(sections["certifications"].weighted, 1),
            "projects_score":   round(sections["projects"].weighted, 1),
            # Raw ratios (0–1) for conditional formatting
            "skills_raw":       round(sections["skills"].raw, 3),
            "experience_raw":   round(sections["experience"].raw, 3),
            "education_raw":    round(sections["education"].raw, 3),
            "summary_raw":      round(sections["summary"].raw, 3),
            # Details
            "years_experience": derived.get("years_experience") or "",
            "education_level":  derived.get("education_level") or "",
            "matched_skills":   ", ".join(matched_skills),
            "missing_skills":   ", ".join(missing_skills),
            "top_gaps":         " | ".join(gaps[:4]),
            "notes":            " | ".join(
                n for s in sections.values() for n in s.notes[:1]
            ),
        }

    def score_folder(
        self,
        folder: Path,
        job: dict,
        workers: int = 1,
    ) -> list[dict[str, Any]]:
        json_files = sorted(folder.glob("*.json"))
        if not json_files:
            raise FileNotFoundError(f"No JSON files found in {folder}")

        print(f"Scoring {len(json_files)} resumes...", flush=True)
        results: list[dict[str, Any]] = []
        failed: list[str] = []

        # NOTE: HDCSearch has a shared cache + lock — workers > 1 is safe
        # but on CPU it won't be faster since encode is the bottleneck.
        # Keep workers=1 unless you have a GPU.
        def _process(path: Path) -> dict[str, Any]:
            raw = path.read_bytes()
            for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
                try:
                    resume = json.loads(raw.decode(enc))
                    break
                except (UnicodeDecodeError, json.JSONDecodeError):
                    continue
            else:
                raise ValueError(f"Could not decode {path.name}")
            return self.score_one(resume, job)

        with ThreadPoolExecutor(max_workers=workers) as pool:
            futures = {pool.submit(_process, p): p for p in json_files}
            done = 0
            for future in as_completed(futures):
                path = futures[future]
                done += 1
                try:
                    result = future.result()
                    results.append(result)
                    print(
                        f"  [{done}/{len(json_files)}] "
                        f"{result['name']:<30} {result['total_score']:5.1f}  {result['grade']}",
                        flush=True,
                    )
                except Exception as e:
                    failed.append(path.name)
                    print(f"  [{done}/{len(json_files)}] FAILED {path.name}: {e}", flush=True)

        if failed:
            print(f"\nFailed files ({len(failed)}): {', '.join(failed)}", flush=True)

        # Sort by score descending
        results.sort(key=lambda r: r["total_score"], reverse=True)
        return results


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

def _hex(h: str) -> str:
    return h.lstrip("#")

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=_hex(hex_color))

def _font(bold=False, color="000000", size=11) -> Font:
    return Font(bold=bold, color=_hex(color), size=size, name="Arial")

def _border() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _score_color(raw: float) -> str:
    """Green → Yellow → Red based on 0–1 raw score."""
    if raw >= 0.70: return "C6EFCE"   # green
    if raw >= 0.50: return "FFEB9C"   # yellow
    return "FFC7CE"                   # red

def _grade_color(grade: str) -> str:
    return {
        "A": "375623", "B": "006100", "C": "9C5700",
        "D": "9C0006", "F": "9C0006",
    }.get(grade, "000000")



def write_excel(results: list[dict], job: dict, output: Path) -> None:
    """
    Create a professional multi-sheet Excel resume matching report.

    Sheets:
        1. Dashboard
        2. Candidate Rankings
        3. Score Breakdown
        4. Skill Analysis
        5. Raw Data
    """

    wb = openpyxl.Workbook()

    # ------------------------------------------------------------------
    # Theme
    # ------------------------------------------------------------------

    NAVY = "1F3864"
    BLUE = "2E75B6"
    LIGHT_BLUE = "D9EAF7"
    VERY_LIGHT_BLUE = "EBF3FB"
    WHITE = "FFFFFF"
    DARK = "1F1F1F"
    GREY = "666666"
    LIGHT_GREY = "F2F2F2"
    BORDER_GREY = "D9E1F2"

    GREEN = "C6EFCE"
    GREEN_TEXT = "006100"
    YELLOW = "FFEB9C"
    YELLOW_TEXT = "9C6500"
    ORANGE = "FCE4D6"
    ORANGE_TEXT = "C65911"
    RED = "FFC7CE"
    RED_TEXT = "9C0006"

    THIN_BORDER = Border(
        left=Side(style="thin", color=BORDER_GREY),
        right=Side(style="thin", color=BORDER_GREY),
        top=Side(style="thin", color=BORDER_GREY),
        bottom=Side(style="thin", color=BORDER_GREY),
    )

    job_title = job.get("title") or "Job Posting"

    # ------------------------------------------------------------------
    # Helper functions
    # ------------------------------------------------------------------

    def title_style(cell):
        cell.font = Font(
            bold=True,
            size=16,
            color=WHITE,
            name="Arial"
        )
        cell.fill = _fill(NAVY)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    def section_header(cell):
        cell.font = Font(
            bold=True,
            size=11,
            color=WHITE,
            name="Arial"
        )
        cell.fill = _fill(BLUE)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )
        cell.border = THIN_BORDER

    def column_header(cell):
        cell.font = Font(
            bold=True,
            size=10,
            color=WHITE,
            name="Arial"
        )
        cell.fill = _fill(NAVY)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )
        cell.border = THIN_BORDER

    def body_cell(cell, wrap=False):
        cell.font = Font(
            size=10,
            color=DARK,
            name="Arial"
        )
        cell.border = THIN_BORDER
        cell.alignment = Alignment(
            vertical="center",
            wrap_text=wrap
        )

    def score_fill(score):
        """
        Score is 0-100.
        """
        if score >= 85:
            return GREEN
        if score >= 70:
            return "E2F0D9"
        if score >= 55:
            return YELLOW
        if score >= 40:
            return ORANGE
        return RED

    def score_text(score):
        if score >= 85:
            return GREEN_TEXT
        if score >= 70:
            return "548235"
        if score >= 55:
            return YELLOW_TEXT
        if score >= 40:
            return ORANGE_TEXT
        return RED_TEXT

    def raw_fill(raw):
        if raw >= 0.70:
            return GREEN
        if raw >= 0.50:
            return YELLOW
        return RED

    def raw_status(raw):
        if raw >= 0.70:
            return "Strong"
        if raw >= 0.50:
            return "Partial"
        return "Weak"

    # ------------------------------------------------------------------
    # SHEET 1: DASHBOARD
    # ------------------------------------------------------------------

    ws = wb.active
    ws.title = "Dashboard"

    ws.merge_cells("A1:H1")
    ws["A1"] = f"Resume Matching Dashboard — {job_title}"
    title_style(ws["A1"])
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:H2")
    ws["A2"] = "Automated resume evaluation and candidate ranking report"
    ws["A2"].font = Font(
        italic=True,
        size=10,
        color=GREY,
        name="Arial"
    )
    ws["A2"].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    # Column widths
    dashboard_widths = {
        "A": 24,
        "B": 18,
        "C": 5,
        "D": 24,
        "E": 18,
        "F": 5,
        "G": 24,
        "H": 18,
    }

    for col, width in dashboard_widths.items():
        ws.column_dimensions[col].width = width

    scores = [r["total_score"] for r in results]
    grades = [r["grade"] for r in results]

    avg_score = sum(scores) / len(scores) if scores else 0
    highest = max(scores) if scores else 0
    lowest = min(scores) if scores else 0

    # Determine recommended candidates
    recommended = [
        r for r in results
        if r["total_score"] >= 70
    ]

    # --------------------------------------------------------------
    # Key metrics
    # --------------------------------------------------------------

    ws.merge_cells("A4:B4")
    ws["A4"] = "JOB OVERVIEW"
    section_header(ws["A4"])

    overview = [
        ("Job Title", job_title),
        ("Resumes Evaluated", len(results)),
        ("Average Score", avg_score),
        ("Highest Score", highest),
        ("Lowest Score", lowest),
        ("Recommended Candidates", len(recommended)),
    ]

    for i, (label, value) in enumerate(overview, 5):
        ws.cell(i, 1, label)
        ws.cell(i, 2, value)

        body_cell(ws.cell(i, 1))
        body_cell(ws.cell(i, 2))

        ws.cell(i, 1).font = Font(
            bold=True,
            size=10,
            name="Arial"
        )

        if isinstance(value, float):
            ws.cell(i, 2).number_format = "0.0"

    # --------------------------------------------------------------
    # Grade distribution
    # --------------------------------------------------------------

    ws.merge_cells("D4:E4")
    ws["D4"] = "GRADE DISTRIBUTION"
    section_header(ws["D4"])

    grade_rows = [
        ("A — Excellent", "A"),
        ("B — Strong", "B"),
        ("C — Moderate", "C"),
        ("D — Weak", "D"),
        ("F — Poor", "F"),
    ]

    for i, (label, grade) in enumerate(grade_rows, 5):
        ws.cell(i, 4, label)
        ws.cell(i, 5, grades.count(grade))

        body_cell(ws.cell(i, 4))
        body_cell(ws.cell(i, 5))

        ws.cell(i, 5).alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    # --------------------------------------------------------------
    # Score distribution
    # --------------------------------------------------------------

    ws.merge_cells("A13:B13")
    ws["A13"] = "SCORE DISTRIBUTION"
    section_header(ws["A13"])

    distribution = [
        ("85–100", lambda x: x >= 85),
        ("70–84", lambda x: 70 <= x < 85),
        ("55–69", lambda x: 55 <= x < 70),
        ("40–54", lambda x: 40 <= x < 55),
        ("0–39", lambda x: x < 40),
    ]

    for i, (label, condition) in enumerate(distribution, 14):
        count = sum(1 for score in scores if condition(score))

        ws.cell(i, 1, label)
        ws.cell(i, 2, count)

        body_cell(ws.cell(i, 1))
        body_cell(ws.cell(i, 2))

        ws.cell(i, 2).alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    # --------------------------------------------------------------
    # Top candidates
    # --------------------------------------------------------------

    ws.merge_cells("D12:H12")
    ws["D12"] = "TOP CANDIDATES"
    section_header(ws["D12"])

    top_headers = [
        "Rank",
        "Candidate",
        "Score",
        "Grade",
        "Recommendation",
    ]

    for col, header in enumerate(top_headers, 4):
        cell = ws.cell(13, col, header)
        column_header(cell)

    for row_num, candidate in enumerate(results[:5], 14):
        values = [
            row_num - 13,
            candidate["name"],
            candidate["total_score"],
            candidate["grade"],
            candidate["recommendation"],
        ]

        for col, value in enumerate(values, 4):
            cell = ws.cell(row_num, col, value)
            body_cell(cell, wrap=(col == 8))

            if col == 6:
                cell.number_format = "0.0"
                cell.fill = _fill(score_fill(candidate["total_score"]))
                cell.font = Font(
                    bold=True,
                    color=score_text(candidate["total_score"]),
                    size=10,
                    name="Arial"
                )

            if col == 7:
                cell.font = Font(
                    bold=True,
                    color=_grade_color(candidate["grade"]),
                    size=10,
                    name="Arial"
                )
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

    # --------------------------------------------------------------
    # Score interpretation
    # --------------------------------------------------------------

    ws.merge_cells("A21:H21")
    ws["A21"] = "SCORE INTERPRETATION"
    section_header(ws["A21"])

    interpretation = [
        ("85–100", "Excellent Match"),
        ("70–84", "Strong Match"),
        ("55–69", "Moderate Match"),
        ("40–54", "Weak Match"),
        ("0–39", "Poor Match"),
    ]

    for i, (score_range, description) in enumerate(interpretation, 22):
        ws.cell(i, 1, score_range)
        ws.cell(i, 2, description)

        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=8)

        body_cell(ws.cell(i, 1))
        body_cell(ws.cell(i, 2))

        score_start = int(score_range.split("–")[0])
        ws.cell(i, 1).fill = _fill(score_fill(score_start))
        ws.cell(i, 1).font = Font(
            bold=True,
            color=score_text(score_start),
            size=10,
            name="Arial"
        )

    # ------------------------------------------------------------------
    # SHEET 2: CANDIDATE RANKINGS
    # ------------------------------------------------------------------

    ws = wb.create_sheet("Candidate Rankings")

    # 1st row: title
    ws.merge_cells("A1:V1")
    ws["A1"] = f"Candidate Rankings — {job_title}"
    title_style(ws["A1"])
    ws.row_dimensions[1].height = 30

    # 2nd row: grouped headers
    groups = [
        ("Candidate Information", 1, 5),
        ("Overall Assessment", 6, 8),
        ("Score Breakdown", 9, 14),
        ("Candidate Profile", 15, 16),
        ("Skills & Gaps", 17, 19),
        ("Source", 20, 20),
    ]

    for group_name, start_col, end_col in groups:
        ws.merge_cells(
            start_row=2,
            start_column=start_col,
            end_row=2,
            end_column=end_col
        )
        cell = ws.cell(2, start_col, group_name)
        section_header(cell)

    # 3rd row: column headers
    ranking_headers = [
        ("Rank", 7),
        ("Candidate", 24),
        ("Email", 28),
        ("Phone", 16),
        ("Location", 20),

        ("Overall Score", 14),
        ("Grade", 9),
        ("Recommendation", 32),

        ("Skills", 12),
        ("Experience", 14),
        ("Education", 13),
        ("Summary", 12),
        ("Certifications", 16),
        ("Projects", 12),

        ("Years Experience", 16),
        ("Education Level", 20),

        ("Matched Skills", 38),
        ("Missing Skills", 38),
        ("Top Gaps", 42),

        ("Source File", 28),
    ]

    for col, (header, width) in enumerate(ranking_headers, 1):
        cell = ws.cell(3, col, header)
        column_header(cell)
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[3].height = 34

    for row_idx, res in enumerate(results, 4):
        values = [
            row_idx - 3,
            res.get("name", "Unknown"),
            res.get("email", ""),
            res.get("phone", ""),
            res.get("location", ""),

            res.get("total_score", 0),
            res.get("grade", ""),
            res.get("recommendation", ""),

            res.get("skills_score", 0),
            res.get("experience_score", 0),
            res.get("education_score", 0),
            res.get("summary_score", 0),
            res.get("certs_score", 0),
            res.get("projects_score", 0),

            res.get("years_experience", ""),
            res.get("education_level", ""),

            res.get("matched_skills", ""),
            res.get("missing_skills", ""),
            res.get("top_gaps", ""),

            res.get("source_file", ""),
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row_idx, col_idx, value)

            body_cell(
                cell,
                wrap=col_idx in (8, 17, 18, 19, 20)
            )

            # Alternating row background
            if row_idx % 2 == 0:
                cell.fill = _fill(VERY_LIGHT_BLUE)

            # Overall score
            if col_idx == 6:
                score = float(value or 0)

                cell.number_format = "0.0"
                cell.fill = _fill(score_fill(score))
                cell.font = Font(
                    bold=True,
                    color=score_text(score),
                    size=10,
                    name="Arial"
                )
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

            # Grade
            elif col_idx == 7:
                cell.font = Font(
                    bold=True,
                    color=_grade_color(str(value)),
                    size=11,
                    name="Arial"
                )
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

            # Section scores
            elif 9 <= col_idx <= 14:
                cell.number_format = '0.0'
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

            # Rank
            elif col_idx == 1:
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )
                cell.font = Font(
                    bold=True,
                    size=10,
                    name="Arial"
                )

        ws.row_dimensions[row_idx].height = 38

    # Freeze headers
    ws.freeze_panes = "A4"

    # Correct filter range
    if results:
        ws.auto_filter.ref = (
            f"A3:T{len(results) + 3}"
        )

    # Add Excel table
    if results:
        from openpyxl.worksheet.table import Table, TableStyleInfo

        table = Table(
            displayName="CandidateRankingTable",
            ref=f"A3:T{len(results) + 3}"
        )

        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=False,
            showColumnStripes=False,
        )

        table.tableStyleInfo = style
        ws.add_table(table)

    # ------------------------------------------------------------------
    # SHEET 3: SCORE BREAKDOWN
    # ------------------------------------------------------------------

    ws = wb.create_sheet("Score Breakdown")

    ws.merge_cells("A1:G1")
    ws["A1"] = f"Detailed Score Breakdown — {job_title}"
    title_style(ws["A1"])
    ws.row_dimensions[1].height = 30

    breakdown_headers = [
        ("Candidate", 25),
        ("Category", 18),
        ("Weight", 12),
        ("Raw Score", 14),
        ("Weighted Score", 18),
        ("Status", 14),
        ("Contribution", 16),
    ]

    for col, (header, width) in enumerate(breakdown_headers, 1):
        cell = ws.cell(3, col, header)
        column_header(cell)
        ws.column_dimensions[get_column_letter(col)].width = width

    category_data = [
        ("Skills", "skills_raw", "skills_score", WEIGHTS.get("skills", 35)),
        ("Experience", "experience_raw", "experience_score", WEIGHTS.get("experience", 20)),
        ("Education", "education_raw", "education_score", WEIGHTS.get("education", 15)),
        ("Summary", "summary_raw", "summary_score", WEIGHTS.get("summary", 15)),
        (
            "Certifications",
            None,
            "certs_score",
            WEIGHTS.get("certifications", 10),
        ),
        (
            "Projects",
            None,
            "projects_score",
            WEIGHTS.get("projects", 5),
        ),
    ]

    row = 4

    for res in results:
        for category, raw_key, weighted_key, weight in category_data:

            weighted_score = float(res.get(weighted_key, 0) or 0)

            if raw_key:
                raw = float(res.get(raw_key, 0) or 0)
            else:
                raw = (
                    weighted_score / weight
                    if weight
                    else 0
                )

            status = raw_status(raw)

            values = [
                res.get("name", "Unknown"),
                category,
                weight,
                raw,
                weighted_score,
                status,
                f"{(weighted_score / res['total_score'] * 100):.1f}%"
                if res.get("total_score")
                else "0%",
            ]

            for col, value in enumerate(values, 1):
                cell = ws.cell(row, col, value)
                body_cell(cell)

                if row % 2 == 0:
                    cell.fill = _fill(VERY_LIGHT_BLUE)

                if col == 3:
                    cell.number_format = '0.0'
                    cell.alignment = Alignment(
                        horizontal="center",
                        vertical="center"
                    )

                elif col == 4:
                    cell.number_format = '0.0%'
                    cell.fill = _fill(raw_fill(raw))
                    cell.alignment = Alignment(
                        horizontal="center",
                        vertical="center"
                    )

                elif col == 5:
                    cell.number_format = '0.0'
                    cell.alignment = Alignment(
                        horizontal="center",
                        vertical="center"
                    )

                elif col == 6:
                    cell.alignment = Alignment(
                        horizontal="center",
                        vertical="center"
                    )

                    if status == "Strong":
                        cell.fill = _fill(GREEN)
                        cell.font = Font(
                            bold=True,
                            color=GREEN_TEXT,
                            name="Arial",
                            size=10
                        )
                    elif status == "Partial":
                        cell.fill = _fill(YELLOW)
                        cell.font = Font(
                            bold=True,
                            color=YELLOW_TEXT,
                            name="Arial",
                            size=10
                        )
                    else:
                        cell.fill = _fill(RED)
                        cell.font = Font(
                            bold=True,
                            color=RED_TEXT,
                            name="Arial",
                            size=10
                        )

            row += 1

    ws.freeze_panes = "A4"

    if row > 4:
        ws.auto_filter.ref = f"A3:G{row - 1}"

    # ------------------------------------------------------------------
    # SHEET 4: SKILL ANALYSIS
    # ------------------------------------------------------------------

    ws = wb.create_sheet("Skill Analysis")

    ws.merge_cells("A1:D1")
    ws["A1"] = f"Skill Analysis — {job_title}"
    title_style(ws["A1"])
    ws.row_dimensions[1].height = 30

    # Extract required skills from job posting
    job_skills = (
        job.get("required_skills")
        or job.get("all_skills")
        or []
    )

    if isinstance(job_skills, str):
        job_skills = [x.strip() for x in job_skills.split(",")]

    skill_headers = [
        ("Required Skill", 30),
        ("Candidates Matching", 22),
        ("Candidates Missing", 22),
        ("Match Rate", 16),
    ]

    for col, (header, width) in enumerate(skill_headers, 1):
        cell = ws.cell(3, col, header)
        column_header(cell)
        ws.column_dimensions[get_column_letter(col)].width = width

    for row_idx, skill in enumerate(job_skills, 4):
        skill = str(skill).strip()

        if not skill:
            continue

        skill_lower = skill.lower()

        matching = 0

        for res in results:
            matched = str(
                res.get("matched_skills", "")
            ).lower()

            if skill_lower in matched:
                matching += 1

        total_candidates = len(results)
        missing = max(total_candidates - matching, 0)

        match_rate = (
            matching / total_candidates
            if total_candidates
            else 0
        )

        values = [
            skill,
            matching,
            missing,
            match_rate,
        ]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row_idx, col, value)
            body_cell(cell)

            if row_idx % 2 == 0:
                cell.fill = _fill(VERY_LIGHT_BLUE)

            if col in (2, 3):
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

            if col == 4:
                cell.number_format = "0.0%"
                cell.fill = _fill(raw_fill(match_rate))
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

    ws.freeze_panes = "A4"

    if job_skills:
        ws.auto_filter.ref = f"A3:D{3 + len(job_skills)}"

    # --------------------------------------------------------------
    # Candidate skill summary
    # --------------------------------------------------------------

    start_row = max(7, 5 + len(job_skills))

    ws.merge_cells(
        start_row=start_row,
        start_column=1,
        end_row=start_row,
        end_column=4
    )
    ws.cell(start_row, 1, "CANDIDATE SKILL SUMMARY")
    section_header(ws.cell(start_row, 1))

    summary_headers = [
        "Candidate",
        "Matched Skills",
        "Missing Skills",
        "Top Gaps",
    ]

    for col, header in enumerate(summary_headers, 1):
        cell = ws.cell(start_row + 1, col, header)
        column_header(cell)

    for idx, res in enumerate(results, start_row + 2):
        values = [
            res.get("name", "Unknown"),
            res.get("matched_skills", ""),
            res.get("missing_skills", ""),
            res.get("top_gaps", ""),
        ]

        for col, value in enumerate(values, 1):
            cell = ws.cell(idx, col, value)
            body_cell(cell, wrap=True)

            if idx % 2 == 0:
                cell.fill = _fill(VERY_LIGHT_BLUE)

        ws.row_dimensions[idx].height = 40

    # ------------------------------------------------------------------
    # SHEET 5: RAW DATA
    # ------------------------------------------------------------------

    ws = wb.create_sheet("Raw Data")

    raw_headers = [
        "rank",
        "name",
        "total_score",
        "grade",
        "skills_score",
        "experience_score",
        "education_score",
        "summary_score",
        "certs_score",
        "projects_score",
        "years_experience",
        "education_level",
        "matched_skills",
        "missing_skills",
        "top_gaps",
        "recommendation",
        "email",
        "phone",
        "location",
        "source_file",
    ]

    for col, header in enumerate(raw_headers, 1):
        cell = ws.cell(1, col, header)
        column_header(cell)

        ws.column_dimensions[
            get_column_letter(col)
        ].width = max(14, min(len(header) + 5, 30))

    for rank, res in enumerate(results, 1):
        values = [
            rank
        ] + [
            res.get(h, "")
            for h in raw_headers[1:]
        ]

        for col, value in enumerate(values, 1):
            cell = ws.cell(rank + 1, col, value)
            body_cell(
                cell,
                wrap=col in (13, 14, 15, 16)
            )

            if rank % 2 == 0:
                cell.fill = _fill(VERY_LIGHT_BLUE)

            if col == 3:
                cell.number_format = "0.0"

    ws.freeze_panes = "A2"

    if results:
        ws.auto_filter.ref = (
            f"A1:T{len(results) + 1}"
        )

    # ------------------------------------------------------------------
    # Global worksheet formatting
    # ------------------------------------------------------------------

    for worksheet in wb.worksheets:
        worksheet.sheet_view.showGridLines = False

        # Set Arial as the default-ish font for populated cells
        for row_cells in worksheet.iter_rows():
            for cell in row_cells:
                if cell.value is not None and cell.font.name != "Arial":
                    cell.font = Font(
                        name="Arial",
                        size=cell.font.sz or 10,
                        bold=cell.font.bold,
                        italic=cell.font.italic,
                        color=cell.font.color,
                    )

    # ------------------------------------------------------------------
    # Add conditional formatting to Candidate Rankings
    # ------------------------------------------------------------------

    from openpyxl.formatting.rule import ColorScaleRule, DataBarRule

    rankings = wb["Candidate Rankings"]

    if results:
        last_row = len(results) + 3

        # Overall score data bar
        rankings.conditional_formatting.add(
            f"F4:F{last_row}",
            DataBarRule(
                start_type="num",
                start_value=0,
                end_type="num",
                end_value=100,
                color="2E75B6",
                showValue=True,
            )
        )

        # Section score color scales
        for col in "IJKLMN":
            rankings.conditional_formatting.add(
                f"{col}4:{col}{last_row}",
                ColorScaleRule(
                    start_type="min",
                    start_color=RED,
                    mid_type="percentile",
                    mid_value=50,
                    mid_color=YELLOW,
                    end_type="max",
                    end_color=GREEN,
                )
            )

    # ------------------------------------------------------------------
    # Save
    # ------------------------------------------------------------------

    wb.save(output)

    print(f"\nSaved professional Excel report: {output}", flush=True)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> None:
    ap = argparse.ArgumentParser(
        description="Score a folder of resume JSONs against one job posting."
    )
    ap.add_argument("resumes_folder", type=Path,
                    help="Folder containing parsed resume .json files")
    ap.add_argument("job_posting",    type=Path,
                    help="Job posting PDF or text file")
    ap.add_argument("--output",  type=Path, default=None,
                    help="Output .xlsx path (default: results.xlsx next to job posting)")
    ap.add_argument("--workers", type=int, default=1,
                    help="Parallel workers (default 1; GPU: try 2–4)")
    args = ap.parse_args()

    if not args.resumes_folder.is_dir():
        ap.error(f"Not a directory: {args.resumes_folder}")
    if not args.job_posting.is_file():
        ap.error(f"Job posting not found: {args.job_posting}")

    output = args.output or args.job_posting.with_name(
        args.job_posting.stem + "_results.xlsx"
    )

    t0      = time.perf_counter()
    scorer  = BatchScorer()
    job     = scorer.load_job(args.job_posting)
    results = scorer.score_folder(args.resumes_folder, job, workers=args.workers)

    if not results:
        ap.error("No resumes were scored successfully.")

    write_excel(results, job, output)
    elapsed = time.perf_counter() - t0
    print(f"Done. {len(results)} resumes in {elapsed:.1f}s "
          f"({elapsed/len(results):.1f}s each)", flush=True)


if __name__ == "__main__":
    main()